from http.client import HTTPException
import os
from fastapi.responses import JSONResponse
from fastapi.responses import HTMLResponse

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from app import app
from openai import OpenAI
import json
from openpyxl import load_workbook


f = open('key.json')
key = json.load(f)["key"]
client = OpenAI(api_key=key)


@app.get("/")
async def root() -> JSONResponse:
    return {"response": "Hello World!"}

# name of journal is
# 10_5_24_J.xlsx
# name of credit card entries
# 10_5_24_CC.xlsx


@app.post("/upload")
async def upload_excel_file(file: UploadFile = File(...), first_name: str = Form(...), month: str = Form(...), year: str = Form(...)) -> JSONResponse:
    # Use the custom name provided by the user
    # if not name.endswith(".xlsx"):
    #     name += ".xlsx"  # Ensure the custom name has the correct file extension
    name = first_name + "_" + month + "_" + year + ".xlsx"

    # Save the file with the custom name in the "data" folder
    file_location = os.path.join("data", name)

    try:
        with open(file_location, "wb") as f:
            while contents := await file.read(1024):
                f.write(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")

    return JSONResponse(content={"message": f"File '{name}' uploaded successfully to '{file_location}'."})


def chat(prompt):
    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {
                "role": "user",
                "content": prompt
            }
        ]
    )

    return str(completion.choices[0].message.content)
@app.post("/analyze")
async def analyze(first_name, month, year):
    name = "data/" + first_name + "_" + month + "_" + year + ".xlsx"
 
    analysis = "<h2>Feelings and Analysis:</h2>"
    print("we analyze and got this: " + chat("how much money do you charge?"))

    results = loadExcelFN(name)
    print("We loaded excel: " + str(results))
    finalDictionaryData  = getRowsByCategorySorted(name, results)
    
    counter = 0
    
    for key, value in finalDictionaryData.items():
        counter += 1
        oneAnalysis = chat(f"This is the users spending history for whenever they felt {key}. History here: {str(value)}. Please analyze and give me your analysis.Very brief 3 sentence reply")
        
        # Format with HTML tags for better readability
        newline = f"""
        <div>
            <h3>{counter}. {key.capitalize()}</h3>
            <hr>
            <p>{oneAnalysis}</p>
        </div>
        """
        print(newline)
        analysis += newline
    
    return HTMLResponse(content=analysis)


def loadExcelFN(fileName):

    workbook = load_workbook(fileName)
    sheet = workbook['Sheet1']

    # Create a dictionary to hold the category and its corresponding row numbers
    category_dict = {}

    # Iterate over the rows, skipping the header
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            # Skip the header row
            continue

        # Assuming 'category' is in the third column (index 2)
        category = row[3]

        # Add the row number (idx) to the category's list in the dictionary
        if category in category_dict:
            category_dict[category].append(idx)
        else:
            category_dict[category] = [idx]
    return category_dict


def getRowsByCategorySorted(fileName, category_dict, threshold=2):
    # Load the workbook and sheet
    workbook = load_workbook(fileName)
    sheet = workbook['Sheet1']
    
    # Create a dictionary to hold the row data for each category
    row_data_by_category = {}

    # Iterate over the categories and their row indices in category_dict
    for category, row_indices in category_dict.items():
        if len(row_indices) < threshold:
            # Skip categories with fewer rows than the threshold
            continue

        row_data_by_category[category] = []

        # For each row index, get the row's data and append it to the category list
        for row_idx in row_indices:
            row_data = list(sheet.iter_rows(values_only=True))[row_idx - 1]
            row_data_by_category[category].append(row_data)

    # Sort the dictionary by the length of the list values in descending order
    sorted_row_data_by_category = dict(
        sorted(row_data_by_category.items(), key=lambda item: len(item[1]), reverse=True)
    )

    return sorted_row_data_by_category


@app.get("/loadExcel")
async def loadExcel(first_name: str = Form(...), month: str = Form(...), year: str = Form(...)):
    name = "data/" + first_name + "_" + month + "_" + year + ".xlsx"

    category_dict = getRowsByCategorySorted(name, loadExcelFN(name))

    return {"response": category_dict}
